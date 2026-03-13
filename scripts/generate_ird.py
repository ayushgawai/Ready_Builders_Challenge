"""
Generate docs/IRD.xlsx — Integration Requirements Document with two sheets.

Run: python scripts/generate_ird.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
SECTION_BG  = "D6E4F0"
HEADER_BG   = "1F3864"
ALT_ROW     = "EBF3FB"
WHITE       = "FFFFFF"
YELLOW      = "FFF2CC"
GREEN       = "E2EFDA"
RED         = "FCE4D6"
GREY        = "F2F2F2"

def _hdr_font(bold=True, size=10, colour=WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=colour)

def _cell_font(bold=False, size=10, colour="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=colour)

def _fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def _wrap():
    return Alignment(wrap_text=True, vertical="top")

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def _section_header(ws, row, text, ncols=4):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill  = _fill(MED_BLUE)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 18

def _table_header(ws, row, columns: list):
    for col_idx, label in enumerate(columns, 1):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.font      = _hdr_font()
        c.fill      = _fill(HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _thin_border()
    ws.row_dimensions[row].height = 24

def _data_row(ws, row, values: list, fill_hex=WHITE, bold_first=False):
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font      = Font(name="Calibri", bold=(bold_first and col_idx == 1), size=10)
        c.fill      = _fill(fill_hex)
        c.alignment = _wrap()
        c.border    = _thin_border()


# ============================================================
# SHEET 1 — Project & Integration Requirements
# ============================================================

def build_sheet1(wb: Workbook):
    ws = wb.active
    ws.title = "Sheet1 – Project Requirements"
    ws.sheet_view.showGridLines = False

    _set_col_widths(ws, {
        "A": 30, "B": 35, "C": 55, "D": 40
    })

    # Title row
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "Integration Requirements Document (IRD)  —  LEO Satellite Coverage Risk Pipeline"
    c.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill      = _fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Sub-title
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value     = "Sheet 1 of 2 — Project & Integration Requirements"
    c.font      = Font(name="Calibri", italic=True, size=10, color="555555")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    row = 4

    # ---------- SECTION: Project Info ----------
    _section_header(ws, row, "PROJECT INFORMATION"); row += 1
    _table_header(ws, row, ["Category", "Item", "Detail", "Notes / Status"]); row += 1

    project_info = [
        ("Project Info",  "Project Name",   "LEO Satellite Coverage Risk Analysis", ""),
        ("",              "Challenge",       "Ready.net Builders Challenge — Issue #50",
         "https://github.com/ready/builders-challenge/issues/50"),
        ("",              "Developer",       "Ayush Gawai", ""),
        ("",              "Version",         "1.0", ""),
        ("",              "Date",            "March 2026", ""),
        ("",              "Status",          "Active Development", ""),
    ]
    for i, row_data in enumerate(project_info):
        _data_row(ws, row, list(row_data), ALT_ROW if i % 2 else WHITE)
        row += 1

    row += 1
    # ---------- SECTION: Contacts ----------
    _section_header(ws, row, "CONTACTS / STAKEHOLDERS"); row += 1
    _table_header(ws, row, ["Role", "Name", "Contact", "Notes"]); row += 1
    contacts = [
        ("Challenge Owner", "Ready.net Engineering Team", "builders-challenge@ready.net", ""),
        ("Developer",       "Ayush Gawai",                "—", ""),
    ]
    for i, r in enumerate(contacts):
        _data_row(ws, row, list(r), ALT_ROW if i % 2 else WHITE)
        row += 1

    row += 1
    # ---------- SECTION: Input Data ----------
    _section_header(ws, row, "INPUT DATA REQUIREMENTS"); row += 1
    _table_header(ws, row, ["Item", "Detail", "Value / Path", "Notes"]); row += 1
    inputs = [
        ("File Name",             "DATA_CHALLENGE_50.csv",
         "Provided by Ready.net team", ""),
        ("Format",                "CSV — comma separated · UTF-8",
         "data/raw/DATA_CHALLENGE_50.csv", "gitignored — must be placed manually"),
        ("Size (observed)",       "4,674,919 rows × 4 columns",
         "~4.67M broadband locations", ""),
        ("Frequency",             "One-time batch load (not streaming)", "", ""),
        ("Columns (actual file)", "location_id · latitude · longitude · geoid_cb",
         "geoid_cb = 15-digit Census Block GEOID",
         "Challenge spec listed state/county — actual file has geoid_cb instead. Pipeline derives both."),
        ("Columns (challenge spec)", "location_id · latitude · longitude · state · county",
         "All 5 confirmed in challenge doc", ""),
    ]
    for i, r in enumerate(inputs):
        _data_row(ws, row, list(r), ALT_ROW if i % 2 else WHITE)
        row += 1

    row += 1
    # ---------- SECTION: External Data ----------
    _section_header(ws, row, "EXTERNAL DATA SOURCES (rasters — auto-downloaded)"); row += 1
    _table_header(ws, row, ["Dataset", "Description", "Resolution / Source", "Notes"]); row += 1
    rasters = [
        ("NLCD Tree Canopy Cover 2021",
         "% of pixel covered by tree canopy",
         "30m · USGS / MRLC",
         "Downloaded once · cached in data/raw/"),
        ("USGS 3DEP DEM",
         "Digital Elevation Model → compute terrain slope",
         "10-30m · USGS National Map",
         ""),
        ("NLCD Land Cover 2021",
         "Land use / land cover classification codes",
         "30m · USGS / MRLC",
         ""),
    ]
    for i, r in enumerate(rasters):
        _data_row(ws, row, list(r), ALT_ROW if i % 2 else WHITE)
        row += 1

    row += 1
    # ---------- SECTION: Output Files ----------
    _section_header(ws, row, "OUTPUT FILES"); row += 1
    _table_header(ws, row, ["File", "Location", "Format", "Description"]); row += 1
    outputs = [
        ("locations_scored.csv",        "data/processed/", "CSV",      "All valid locations with risk tier + component scores — main deliverable"),
        ("data_quality_report.txt",     "data/output/",    "Text",     "Drop counts per reason · retention rate"),
        ("anomaly_report.txt",          "data/output/",    "Text",     "Validation anomalies · impossible values · distribution checks"),
        ("findings_report.md",          "data/output/",    "Markdown", "Summary statistics and key findings narrative"),
        ("risk_distribution.png",       "data/output/",    "PNG",      "Bar / pie chart of HIGH / MODERATE / LOW distribution"),
        ("risk_map_static.png",         "data/output/",    "PNG",      "Static map coloured by risk tier"),
        ("risk_map_interactive.html",   "data/output/",    "HTML",     "Interactive Folium map with hover tooltips (bonus)"),
    ]
    for i, r in enumerate(outputs):
        _data_row(ws, row, list(r), ALT_ROW if i % 2 else WHITE)
        row += 1

    row += 1
    # ---------- SECTION: API / Credentials ----------
    _section_header(ws, row, "API / CREDENTIALS"); row += 1
    _table_header(ws, row, ["Item", "Detail", "Location", "Status"]); row += 1
    creds = [
        ("Anthropic API Key",
         "Required for agent orchestration (Claude claude-opus-4-5)",
         ".env — ANTHROPIC_API_KEY=…",
         "PENDING — provided key had zero credit balance. Emailed team."),
        ("Fallback (no key)",
         "python -m src.main --mode batch --no-agent",
         "Direct function calls — no LLM cost",
         "Available"),
    ]
    for i, r in enumerate(creds):
        bg = YELLOW if i == 0 else (ALT_ROW if i % 2 else WHITE)
        _data_row(ws, row, list(r), bg)
        row += 1

    row += 1
    # ---------- SECTION: Tech Stack ----------
    _section_header(ws, row, "TECHNOLOGY STACK"); row += 1
    _table_header(ws, row, ["Component", "Technology", "Version", "Notes"]); row += 1
    stack = [
        ("Language",                "Python",                      "3.11+",        ""),
        ("AI Agent / Orchestration","Anthropic Claude API (tool_use)","claude-opus-4-5","Swap to claude-haiku-4-5 for cheaper dev runs"),
        ("Raster I/O",              "rasterio",                    "≥1.3",         ""),
        ("Spatial / vector",        "geopandas · shapely",         "—",            ""),
        ("Tabular data",            "pandas · numpy",              "≥2.0",         ""),
        ("Reverse geocoding",       "reverse_geocoder",            "offline KD-tree","No API key · no network calls at runtime"),
        ("Visualization",           "matplotlib · folium",         "—",            ""),
        ("Testing",                 "pytest",                      "≥8.0",         ""),
    ]
    for i, r in enumerate(stack):
        _data_row(ws, row, list(r), ALT_ROW if i % 2 else WHITE)
        row += 1

    row += 1
    # ---------- SECTION: Idempotency ----------
    _section_header(ws, row, "IDEMPOTENCY / CACHING"); row += 1
    _table_header(ws, row, ["Aspect", "Detail", "Implementation", "Notes"]); row += 1
    cache = [
        ("Behaviour",        "Pipeline processes data ONCE and caches the scored output",
         "is_scored_cache_valid() in src/utils/pipeline_utils.py",
         "Avoids 20-60 min raster sampling on re-runs"),
        ("Cache check",      "Compares file modification time: output.mtime ≥ input.mtime",
         "Instantaneous — no hashing",
         "Production would use content hash or version manifest"),
        ("Force re-run",     "Delete data/processed/locations_scored.csv",
         "—", ""),
        ("Cache miss action","Run full ingest → enrich → score pipeline, then save cache",
         "save_scored_cache(df, SCORED_LOCATIONS_PATH)", ""),
        ("Cache hit action", "Load scored CSV directly — skip raster sampling",
         "load_scored_cache(SCORED_LOCATIONS_PATH)", ""),
    ]
    for i, r in enumerate(cache):
        _data_row(ws, row, list(r), ALT_ROW if i % 2 else WHITE)
        row += 1

    row += 1
    # ---------- SECTION: Open Items ----------
    _section_header(ws, row, "OPEN ITEMS / MISSING INFORMATION"); row += 1
    _table_header(ws, row, ["Item", "Status", "Action Taken", "Owner"]); row += 1
    open_items = [
        ("Anthropic API credits (key has zero balance)",
         "PENDING", "Emailed Ready.net team to top up account", "Ready.net"),
        ("Output format confirmation",
         "RESOLVED", "Inferred from spec: CSV + Markdown + charts", "Developer"),
        ("Output delivery method",
         "ASSUMED", "GitHub push — no specific format required", "Developer"),
    ]
    colours = [YELLOW, GREEN, GREEN]
    for r_data, bg in zip(open_items, colours):
        _data_row(ws, row, list(r_data), bg)
        row += 1


# ============================================================
# SHEET 2 — Data Dictionary
# ============================================================

def build_sheet2(wb: Workbook):
    ws = wb.create_sheet("Sheet2 – Data Dictionary")
    ws.sheet_view.showGridLines = False

    _set_col_widths(ws, {
        "A": 14, "B": 22, "C": 22, "D": 14,
        "E": 20, "F": 12, "G": 10, "H": 40,
        "I": 18, "J": 38, "K": 45,
    })

    # Title
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value     = "Integration Requirements Document (IRD)  —  Data Dictionary"
    c.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill      = _fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value     = "Sheet 2 of 2 — All fields: Input · Derived · Environmental · Scoring · Quality Checks"
    c.font      = Font(name="Calibri", italic=True, size=10, color="555555")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    COLUMNS = [
        "Sheet / Stage", "Field Name", "Source File",
        "Type", "Format / Valid Values", "Max Length",
        "Nullable", "Validation Rule", "Example",
        "Issues Found in Data", "Description",
    ]

    # Section colour map
    STAGE_FILL = {
        "INPUT":       "DEEBF7",
        "DERIVED":     "E2EFDA",
        "ENVIRONMENT": "FFF2CC",
        "SCORING":     "FCE4D6",
        "QUALITY":     "F2F2F2",
        "ANOMALY":     "F2F2F2",
        "NLCD REF":    "EAF0FB",
    }

    row = 4

    sections = [
        # ---- INPUT fields ----
        ("INPUT", [
            ("INPUT", "location_id", "DATA_CHALLENGE_50.csv",
             "String", "Alphanumeric", "100 chars", "No (critical)",
             "Not null · not whitespace-only · unique (keep first on dup)",
             "LOC_0001",
             "Duplicate IDs found — kept first occurrence. Whitespace-only IDs dropped.",
             "Unique identifier for each broadband location. Primary key for all downstream joins."),

            ("INPUT", "latitude", "DATA_CHALLENGE_50.csv",
             "Float", "WGS84 decimal degrees", "—", "No (critical)",
             "Numeric · 24.396308 ≤ lat ≤ 49.384358 (CONUS bounds)",
             "47.6062",
             "Non-numeric strings possible (e.g. 'N/A') — dropped: non_numeric_latitude. Out-of-range dropped.",
             "Geographic latitude. Used for raster sampling and reverse geocoding. Strings coerced to float; unconvertible rows dropped."),

            ("INPUT", "longitude", "DATA_CHALLENGE_50.csv",
             "Float", "WGS84 decimal degrees", "—", "No (critical)",
             "Numeric · −124.848974 ≤ lon ≤ −66.885444 (CONUS bounds)",
             "-122.3321",
             "Non-numeric strings possible — dropped: non_numeric_longitude.",
             "Geographic longitude. Same handling as latitude."),

            ("INPUT", "geoid_cb", "DATA_CHALLENGE_50.csv",
             "String", "15-digit Census Block GEOID", "15 chars", "Yes (optional)",
             "All digits if present · ≥12 characters",
             "530330101001001",
             "Null values present in some rows. Float64 dtype inference bug (leading zeros stripped) fixed with dtype={'geoid_cb': str}.",
             "Census Block GEOID — SSCCCTTTTTTBBBB (SS=state, CCC=county, TTTTTT=tract, BBBB=block). Used to derive state + county."),
        ]),
        # ---- DERIVED ----
        ("DERIVED", [
            ("DERIVED", "state", "Derived from geoid_cb (FIPS lookup) / reverse geocoder (fallback)",
             "String", "2-letter US state abbreviation", "2 chars", "Yes",
             "Must be in 50-state + DC set if present",
             "WA",
             "Null geoid_cb rows fall back to reverse geocoding. Rows with invalid state are dropped.",
             "Derived: us.states.lookup(state_fips_2char).abbr via the 'us' library. Fallback: us.states.lookup(admin1_full_name).abbr from reverse_geocoder."),

            ("DERIVED", "county", "Census TIGER NAMELSAD via pygris (primary); reverse_geocoder admin2 as-is (fallback)",
             "String", "NAMELSAD, e.g. 'Santa Clara County'", "≤ 60 chars", "Yes",
             "—",
             "Santa Clara County",
             "Null only for rows with both null geoid_cb AND null/invalid coordinates.",
             "Official Census NAMELSAD field — includes legal area designation suffix "
             "('County', 'Parish', 'Borough', etc.). Source: pygris fetches Census TIGER/Line "
             "cartographic boundary (~2MB, cached to ~/.pygris/). Fallback rows use "
             "reverse_geocoder admin2 field as-is (same format)."),

            ("DERIVED", "county_fips", "Derived from geoid_cb (first 5 digits)",
             "String", "5-digit state+county FIPS", "5 chars", "Yes",
             "—",
             "06075",
             "Leading zeros preserved only when geoid_cb read as str dtype. Null if geoid_cb is absent.",
             "5-digit state+county FIPS code (e.g. '06075' = San Francisco County, CA). "
             "Kept as internal column for FCC/BEAD broadband data joins. "
             "Not the same as `county` (name) — this is the government identifier code."),
        ]),
        # ---- ENVIRONMENT ----
        ("ENVIRONMENT", [
            ("ENVIRONMENT", "canopy_pct", "NLCD Tree Canopy Cover 2021 (30m GeoTIFF)",
             "Float", "0.0 – 100.0 (% cover)", "—", "Yes (NaN = raster nodata)",
             "0.0 ≤ value ≤ 100.0",
             "75.0",
             "NaN expected near raster edges / water bodies.",
             "Tree canopy cover % at 30m pixel. Higher = more obstruction risk for satellite dish."),

            ("ENVIRONMENT", "elevation_m", "USGS 3DEP DEM (10-30m GeoTIFF)",
             "Float", "Metres above sea level", "—", "Yes (NaN = raster nodata)",
             "—",
             "1250.5",
             "—",
             "Elevation in metres. Used to compute terrain slope."),

            ("ENVIRONMENT", "slope_deg", "Derived from DEM via central-difference gradient",
             "Float", "0.0 – 90.0 degrees", "—", "Yes (NaN if elevation NaN)",
             "0.0 ≤ value ≤ 90.0",
             "18.3",
             "Values outside range flagged as impossible in anomaly check.",
             "Terrain slope in degrees. Computed with geographic-aware cell size conversion. Steeper = more horizon obstruction."),

            ("ENVIRONMENT", "land_cover_code", "NLCD Land Cover 2021 (30m GeoTIFF)",
             "Integer", "NLCD code (see NLCD REF tab below)", "—", "Yes (NaN = nodata)",
             "Must be a valid NLCD code if present",
             "41",
             "—",
             "NLCD land cover class at the location. Forest codes (41/42/43) = HIGH risk."),
        ]),
        # ---- SCORING ----
        ("SCORING", [
            ("SCORING", "canopy_score", "Computed from canopy_pct",
             "Float", "0.0 – 1.0", "—", "Yes (NaN if canopy_pct = NaN)",
             ">50% → 1.0 (HIGH) · 20–50% → 0.5 (MODERATE) · <20% → 0.0 (LOW)",
             "1.0",
             "—",
             "Normalised canopy risk score. Threshold: Starlink FOV is 100-110° → >50% canopy statistically intersects the sky cone."),

            ("SCORING", "slope_score", "Computed from slope_deg",
             "Float", "0.0 – 1.0", "—", "Yes (NaN if slope_deg = NaN)",
             ">20° → 1.0 · 10–20° → 0.5 · <10° → 0.0",
             "0.5",
             "—",
             "Normalised terrain slope score. >20° creates significant horizon elevation from nearby ridges."),

            ("SCORING", "landcover_score", "Computed from land_cover_code",
             "Float", "0.0 – 1.0", "—", "Yes (NaN if land_cover_code = NaN)",
             "Forest (41/42/43) → 1.0 · Developed (21–24) → 0.5 · Open → 0.0",
             "1.0",
             "—",
             "Normalised land cover obstruction score."),

            ("SCORING", "composite_score",
             "Weighted sum: canopy×0.50 + slope×0.30 + landcover×0.20",
             "Float", "0.0 – 1.0", "—", "Yes (NaN if any input NaN)",
             "Weights sum to 1.0 · NaN propagates",
             "0.75",
             "—",
             "Final composite risk score. NaN propagates because a missing factor = cannot reliably score."),

            ("SCORING", "risk_tier", "Assigned from composite_score",
             "String", "HIGH / MODERATE / LOW / UNKNOWN", "—", "No",
             "≥0.60 → HIGH · 0.30–0.60 → MODERATE · <0.30 → LOW · NaN → UNKNOWN",
             "HIGH",
             "—",
             "Risk classification. UNKNOWN = raster nodata — location validated but not scoreable."),
        ]),
        # ---- QUALITY REPORT ----
        ("QUALITY", [
            ("QUALITY REPORT\n(drop reasons)", "non_numeric_latitude", "data_quality_report.txt",
             "Integer", "Count", "—", "—",
             "Values like 'N/A' · 'abc' · '' in latitude column",
             "3",
             "Indicates source system uses string null sentinel",
             "Rows dropped in Pass 0 because latitude could not be coerced to float."),

            ("", "non_numeric_longitude", "data_quality_report.txt",
             "Integer", "Count", "—", "—",
             "Same as above for longitude",
             "1", "—", ""),

            ("", "null_location_id", "data_quality_report.txt",
             "Integer", "Count", "—", "—", "Null / missing location_id",
             "2", "—", ""),

            ("", "blank_location_id", "data_quality_report.txt",
             "Integer", "Count", "—", "—",
             "Whitespace-only location_id (e.g. '   ')",
             "1", "Whitespace IDs are not null — require explicit strip check", ""),

            ("", "null_latitude", "data_quality_report.txt",
             "Integer", "Count", "—", "—", "Missing latitude", "5", "—", ""),

            ("", "null_longitude", "data_quality_report.txt",
             "Integer", "Count", "—", "—", "Missing longitude", "3", "—", ""),

            ("", "out_of_range_coordinates", "data_quality_report.txt",
             "Integer", "Count", "—", "—",
             "Coords outside CONUS bbox (lat 24.4–49.4 · lon −124.8 to −66.9)",
             "12", "—", ""),

            ("", "duplicate_location_id", "data_quality_report.txt",
             "Integer", "Count", "—", "—",
             "Duplicate location_id — first kept / rest dropped",
             "8", "—", ""),

            ("", "invalid_state_code", "data_quality_report.txt",
             "Integer", "Count", "—", "—",
             "Derived state not in 50-state + DC set",
             "0", "—", ""),
        ]),
        # ---- ANOMALY REPORT ----
        ("ANOMALY", [
            ("ANOMALY REPORT\n(validation checks)", "UNKNOWN tier > 50%", "anomaly_report.txt",
             "Boolean", "CRITICAL", "—", "—",
             ">50% of scored rows are UNKNOWN tier",
             "False",
             "Indicates raster sampling failure (partial download / coverage gap)",
             "Agent halts pipeline if this triggers."),

            ("", "Tier dominance ≥ 90%", "anomaly_report.txt",
             "Boolean", "CRITICAL", "—", "—",
             "One tier accounts for ≥90% of scored rows",
             "False",
             "Indicates scoring bug or wrong raster file loaded",
             "Statistically implausible for a nationally distributed dataset."),

            ("", "Impossible values", "Integer", "CRITICAL",
             "canopy>100 · canopy<0 · slope<0 · slope>90",
             "—", "—", "0", "—",
             "Indicates raster data corruption or unit conversion error.",
             "Any count > 0 triggers CRITICAL."),

            ("", "Forest code + canopy < 5%", "anomaly_report.txt",
             "Integer", "WARNING", "—", "—",
             "NLCD Forest (41/42/43) but canopy_pct < 5%",
             "120",
             "Known NLCD classification lag (recently logged areas)",
             "Flagged for review — NOT dropped. Does not invalidate the run."),
        ]),
        # ---- NLCD REFERENCE ----
        ("NLCD REF", [
            ("NLCD REF", "11 – Open Water", "NLCD", "—", "—", "—", "—", "—", "—", "—", "N/A — not scored"),
            ("", "21 – Developed Open Space",  "NLCD","—","—","—","—","—","—","—","Low-Moderate risk"),
            ("", "22 – Developed Low",         "NLCD","—","—","—","—","—","—","—","Moderate risk"),
            ("", "23 – Developed Medium",       "NLCD","—","—","—","—","—","—","—","Moderate risk"),
            ("", "24 – Developed High",         "NLCD","—","—","—","—","—","—","—","High (dense buildings)"),
            ("", "31 – Barren Land",            "NLCD","—","—","—","—","—","—","—","Low risk"),
            ("", "41 – Deciduous Forest",       "NLCD","—","—","—","—","—","—","—","HIGH — landcover_score = 1.0"),
            ("", "42 – Evergreen Forest",       "NLCD","—","—","—","—","—","—","—","HIGH — landcover_score = 1.0"),
            ("", "43 – Mixed Forest",           "NLCD","—","—","—","—","—","—","—","HIGH — landcover_score = 1.0"),
            ("", "52 – Shrub/Scrub",            "NLCD","—","—","—","—","—","—","—","Low risk"),
            ("", "71 – Grassland/Herbaceous",   "NLCD","—","—","—","—","—","—","—","Low risk"),
            ("", "81 – Pasture/Hay",            "NLCD","—","—","—","—","—","—","—","Low risk"),
            ("", "82 – Cultivated Crops",       "NLCD","—","—","—","—","—","—","—","Low risk"),
        ]),
    ]

    for stage, rows in sections:
        stage_fill = STAGE_FILL.get(stage, WHITE)

        # Section header banner
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=len(COLUMNS)
        )
        c = ws.cell(row=row, column=1, value=stage)
        c.font  = Font(name="Calibri", bold=True, size=11, color=WHITE)
        c.fill  = _fill(MED_BLUE)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 18
        row += 1

        # Column headers
        _table_header(ws, row, COLUMNS)
        row += 1

        for i, r in enumerate(rows):
            bg = stage_fill if i % 2 == 0 else WHITE
            _data_row(ws, row, list(r), bg)
            ws.row_dimensions[row].height = 50
            row += 1

        row += 1  # blank gap between sections

    # Freeze panes on header row of first section
    ws.freeze_panes = "A4"


# ============================================================
# Main
# ============================================================

if __name__ == "__main__":
    out_path = Path(__file__).parent.parent / "docs" / "IRD.xlsx"

    wb = Workbook()
    build_sheet1(wb)
    build_sheet2(wb)

    wb.save(out_path)
    print(f"IRD saved to: {out_path}")
